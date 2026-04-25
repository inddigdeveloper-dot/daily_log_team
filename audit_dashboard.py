import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ========= CONFIG =========
DB_HOST = "localhost"
DB_USER = "root"
DB_PASS = ""                 # XAMPP default is empty; Laragon usually "root"
DB_NAME = "audit_db"

# IMPORTANT: set this to the ACTUAL table name after import
# (check phpMyAdmin — might be "clients" or "clients_social")
CLIENTS_TABLE = "clients"

# Month/year to audit
MONTH = 4
YEAR  = 2026

OUTPUT_FILE = f"audit_{YEAR}_{MONTH:02d}.xlsx"
# ==========================

conn = pymysql.connect(
    host=DB_HOST, user=DB_USER, password=DB_PASS,
    database=DB_NAME, cursorclass=pymysql.cursors.DictCursor
)

# Main query — mirrors the PHP dashboard exactly,
# but WITHOUT the "end_date >= CURDATE()" filter so we can audit expired clients too.
# Flip include_expired = False if you want to match the live dashboard 1:1.
include_expired = True
expiry_filter = "" if include_expired else "WHERE c.end_date >= CURDATE()"

sql = f"""
SELECT 
    c.id,
    c.client_name,
    c.brand_name,
    c.access_received,
    c.end_date,
    DATEDIFF(c.end_date, CURDATE()) AS days_left,
    IFNULL(a.total_posts, 0)     AS total_posts,
    IFNULL(u.posts_done, 0)      AS posts_done,
    IFNULL(u.posts_uploaded, 0)  AS posts_uploaded,
    IFNULL(a.total_reels, 0)     AS total_reels,
    IFNULL(u.reels_done, 0)      AS reels_done,
    IFNULL(u.reels_uploaded, 0)  AS reels_uploaded,
    a.client_id AS alloc_exists,
    u.client_id AS usage_exists
FROM {CLIENTS_TABLE} c
LEFT JOIN allocations a 
  ON c.id = a.client_id AND a.year = %s AND a.month = %s
LEFT JOIN usage_monthly u 
  ON c.id = u.client_id AND u.year = %s AND u.month = %s
{expiry_filter}
ORDER BY c.client_name ASC
"""

with conn.cursor() as cur:
    cur.execute(sql, (YEAR, MONTH, YEAR, MONTH))
    rows = cur.fetchall()

# Also check for orphan records — allocations / usage rows with no matching client
with conn.cursor() as cur:
    cur.execute(f"""
        SELECT a.* FROM allocations a
        LEFT JOIN {CLIENTS_TABLE} c ON a.client_id = c.id
        WHERE c.id IS NULL AND a.year = %s AND a.month = %s
    """, (YEAR, MONTH))
    orphan_allocations = cur.fetchall()

    cur.execute(f"""
        SELECT u.* FROM usage_monthly u
        LEFT JOIN {CLIENTS_TABLE} c ON u.client_id = c.id
        WHERE c.id IS NULL AND u.year = %s AND u.month = %s
    """, (YEAR, MONTH))
    orphan_usage = cur.fetchall()

    # Duplicate allocations/usage for same client+month+year
    cur.execute("""
        SELECT client_id, year, month, COUNT(*) as cnt
        FROM allocations
        WHERE year = %s AND month = %s
        GROUP BY client_id, year, month
        HAVING cnt > 1
    """, (YEAR, MONTH))
    dup_allocations = cur.fetchall()

    cur.execute("""
        SELECT client_id, year, month, COUNT(*) as cnt
        FROM usage_monthly
        WHERE year = %s AND month = %s
        GROUP BY client_id, year, month
        HAVING cnt > 1
    """, (YEAR, MONTH))
    dup_usage = cur.fetchall()

conn.close()

# ========= BUILD EXCEL =========
wb = Workbook()

# ---- Styles ----
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="4E54C8")
issue_fill  = PatternFill("solid", fgColor="FFE0E0")  # light red for flagged rows
thin        = Side(border_style="thin", color="CCCCCC")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)
center      = Alignment(horizontal="center", vertical="center")

def style_headers(ws, headers, row=1):
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

def auto_width(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------- Sheet 1: Dashboard View ----------
ws1 = wb.active
ws1.title = "Dashboard"

headers1 = [
    "Client ID", "Business Name", "Access",
    "Created Posts", "Total Posts", "Uploaded Posts",
    "Created Reels", "Total Reels", "Uploaded Reels",
    "End Date", "Days Left", "Issues"
]
style_headers(ws1, headers1)

total_created_posts = total_uploaded_posts = 0
total_created_reels = total_uploaded_reels = 0
flagged_count = 0

for i, r in enumerate(rows, start=2):
    # --- Audit checks ---
    issues = []
    if r['alloc_exists'] is None:
        issues.append("No allocation row")
    if r['usage_exists'] is None:
        issues.append("No usage row")
    if r['posts_done'] > r['total_posts'] and r['total_posts'] > 0:
        issues.append("posts_done > total_posts")
    if r['reels_done'] > r['total_reels'] and r['total_reels'] > 0:
        issues.append("reels_done > total_reels")
    if r['posts_uploaded'] > r['posts_done']:
        issues.append("uploaded > created (posts)")
    if r['reels_uploaded'] > r['reels_done']:
        issues.append("uploaded > created (reels)")
    if r['access_received'] == 0 and (r['posts_done'] > 0 or r['reels_done'] > 0):
        issues.append("Work done but access pending")

    total_created_posts   += r['posts_done']
    total_uploaded_posts  += r['posts_uploaded']
    total_created_reels   += r['reels_done']
    total_uploaded_reels  += r['reels_uploaded']

    values = [
        r['id'],
        r['brand_name'] or r['client_name'],
        "Active" if r['access_received'] == 1 else "Pending",
        r['posts_done'],
        r['total_posts'],
        r['posts_uploaded'],
        r['reels_done'],
        r['total_reels'],
        r['reels_uploaded'],
        str(r['end_date']),
        r['days_left'],
        "; ".join(issues) if issues else ""
    ]

    for col_idx, val in enumerate(values, start=1):
        c = ws1.cell(row=i, column=col_idx, value=val)
        c.border = border
        c.alignment = center
        if issues:
            c.fill = issue_fill

    if issues:
        flagged_count += 1

# Totals row
total_row = len(rows) + 2
ws1.cell(row=total_row, column=3, value="TOTALS").font = Font(bold=True)
ws1.cell(row=total_row, column=4, value=total_created_posts).font = Font(bold=True)
ws1.cell(row=total_row, column=6, value=total_uploaded_posts).font = Font(bold=True)
ws1.cell(row=total_row, column=7, value=total_created_reels).font = Font(bold=True)
ws1.cell(row=total_row, column=9, value=total_uploaded_reels).font = Font(bold=True)

auto_width(ws1, [10, 35, 12, 14, 12, 14, 14, 12, 14, 14, 11, 45])
ws1.freeze_panes = "A2"

# ---------- Sheet 2: Audit Summary ----------
ws2 = wb.create_sheet("Audit Summary")
summary = [
    ["Audit Period", f"{date(YEAR, MONTH, 1).strftime('%B')} {YEAR}"],
    ["Total Clients", len(rows)],
    ["Clients with Issues", flagged_count],
    ["", ""],
    ["Totals (from data)", ""],
    ["Created Posts", total_created_posts],
    ["Uploaded Posts", total_uploaded_posts],
    ["Created Reels", total_created_reels],
    ["Uploaded Reels", total_uploaded_reels],
    ["", ""],
    ["Data Integrity Issues", ""],
    ["Orphan allocations (no matching client)", len(orphan_allocations)],
    ["Orphan usage records (no matching client)", len(orphan_usage)],
    ["Duplicate allocations (same client+month)", len(dup_allocations)],
    ["Duplicate usage records (same client+month)", len(dup_usage)],
]
for i, row in enumerate(summary, start=1):
    ws2.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
    ws2.cell(row=i, column=2, value=row[1])
auto_width(ws2, [45, 25])

# ---------- Sheet 3: Orphan & Duplicate details ----------
ws3 = wb.create_sheet("Orphans & Duplicates")
r = 1
for label, data in [
    ("Orphan Allocations", orphan_allocations),
    ("Orphan Usage Records", orphan_usage),
    ("Duplicate Allocations", dup_allocations),
    ("Duplicate Usage Records", dup_usage),
]:
    ws3.cell(row=r, column=1, value=label).font = Font(bold=True, size=12)
    r += 1
    if data:
        cols = list(data[0].keys())
        for ci, h in enumerate(cols, start=1):
            c = ws3.cell(row=r, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill
        r += 1
        for rec in data:
            for ci, key in enumerate(cols, start=1):
                ws3.cell(row=r, column=ci, value=str(rec[key]))
            r += 1
    else:
        ws3.cell(row=r, column=1, value="None found ✓")
        r += 1
    r += 2  # spacer

auto_width(ws3, [20, 15, 10, 10, 15, 15, 15, 15])

wb.save(OUTPUT_FILE)

print(f"✅ Audit exported: {OUTPUT_FILE}")
print(f"   Clients audited: {len(rows)}")
print(f"   Clients with issues: {flagged_count}")
print(f"   Orphan allocations: {len(orphan_allocations)}")
print(f"   Orphan usage:       {len(orphan_usage)}")
print(f"   Duplicate allocations: {len(dup_allocations)}")
print(f"   Duplicate usage:       {len(dup_usage)}")
print(f"\n   Totals — Created Posts: {total_created_posts}, Uploaded Posts: {total_uploaded_posts}")
print(f"   Totals — Created Reels: {total_created_reels}, Uploaded Reels: {total_uploaded_reels}")